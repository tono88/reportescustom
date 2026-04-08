import base64
import csv
import io
from datetime import date, datetime, time

from dateutil import parser as date_parser
from markupsafe import escape

from odoo import Command, _, fields, models
from odoo.exceptions import UserError, ValidationError


BASE_HEADER_ALIASES = {
    'product': {'product', 'producto', 'product name'},
    'description': {'description', 'descripcion', 'descripción', 'name', 'detalle'},
    'quantity': {'quantity', 'qty', 'cantidad', 'product_qty'},
    'uom': {'uom', 'unit of measure', 'unidad de medida', 'udm', 'product_uom'},
    'price_unit': {'unit price', 'price', 'precio unitario', 'precio', 'price_unit'},
    'taxes': {'taxes', 'impuestos', 'tax', 'tax ids', 'taxes_id'},
    'date_planned': {
        'delivery date',
        'fecha entrega',
        'fecha de entrega',
        'expected arrival',
        'date planned',
        'date_planned',
        'planned date',
    },
}


class PurchaseOrderLineImportWizard(models.TransientModel):
    _name = 'purchase.order.line.import.wizard'
    _description = 'Import Purchase Order Lines Wizard'

    purchase_id = fields.Many2one('purchase.order', required=True, readonly=True)
    import_file = fields.Binary(string='Archivo', required=False)
    file_name = fields.Char(string='Nombre de archivo')
    has_header = fields.Boolean(string='El archivo tiene encabezados', default=True)
    delimiter = fields.Selection(
        [('comma', 'Coma (,)'), ('semicolon', 'Punto y coma (;)')],
        string='Separador CSV',
        default='comma',
        required=True,
    )
    product_lookup = fields.Selection(
        [
            ('auto', 'Automático: referencia interna, barcode y luego nombre'),
            ('default_code', 'Solo referencia interna'),
            ('barcode', 'Solo código de barras'),
            ('name', 'Solo nombre'),
        ],
        string='Búsqueda de producto',
        default='auto',
        required=True,
    )
    import_mode = fields.Selection(
        [('append', 'Agregar líneas'), ('replace', 'Reemplazar líneas existentes')],
        string='Modo de importación',
        default='append',
        required=True,
    )
    state = fields.Selection([('upload', 'upload'), ('result', 'result')], default='upload')
    result_message = fields.Html(string='Resultado', sanitize=False, readonly=True)

    def action_import(self):
        self.ensure_one()
        if self.purchase_id.state not in ('draft', 'sent'):
            raise UserError(_('Solo se permite importar líneas en RFQ u órdenes en borrador.'))
        if not self.import_file:
            raise UserError(_('Debes adjuntar un archivo CSV o XLSX.'))

        rows, headers = self._read_file()
        if not rows:
            raise UserError(_('El archivo no contiene filas de datos para importar.'))

        if self.import_mode == 'replace':
            self.purchase_id.order_line.filtered(lambda l: not l.display_type).unlink()

        created_count = 0
        skipped = []

        start_index = 2 if self.has_header else 1
        for row_number, row in enumerate(rows, start=start_index):
            if self._row_is_empty(row):
                continue
            with self.env.cr.savepoint():
                try:
                    vals = self._prepare_line_vals(row, headers=headers)
                    self.env['purchase.order.line'].create(vals)
                    created_count += 1
                except Exception as err:
                    skipped.append((row_number, str(err)))

        result_html = self._build_result_html(created_count, skipped)
        self.result_message = result_html
        self.state = 'result'
        self.purchase_id.message_post(body=result_html)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Resultado importación'),
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_close(self):
        return {'type': 'ir.actions.act_window_close'}

    def _read_file(self):
        decoded = base64.b64decode(self.import_file)
        filename = (self.file_name or '').lower()
        if filename.endswith('.csv'):
            return self._read_csv(decoded)
        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            return self._read_xlsx(decoded)
        raise ValidationError(_('Formato no soportado. Usa .csv o .xlsx'))

    def _read_csv(self, decoded):
        delimiter = ',' if self.delimiter == 'comma' else ';'
        try:
            content = decoded.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = decoded.decode('latin-1')
        rows = list(csv.reader(io.StringIO(content), delimiter=delimiter))
        return self._split_headers(rows)

    def _read_xlsx(self, decoded):
        try:
            from openpyxl import load_workbook
        except ImportError as err:
            raise UserError(_('Falta la librería de Python openpyxl en el servidor.')) from err

        workbook = load_workbook(filename=io.BytesIO(decoded), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        return self._split_headers(rows)

    def _split_headers(self, rows):
        clean_rows = [self._clean_row(row) for row in rows]
        if not clean_rows:
            return [], []
        if self.has_header:
            headers = [str(v).strip() if v not in (None, False) else '' for v in clean_rows[0]]
            data_rows = clean_rows[1:]
            return data_rows, headers
        return clean_rows, []

    def _clean_row(self, row):
        values = list(row or [])
        while values and values[-1] in (None, ''):
            values.pop()
        return values

    def _row_is_empty(self, row):
        return all(value in (None, '', False) for value in row)

    def _prepare_line_vals(self, row, headers=None):
        headers = headers or []
        if headers:
            data, extra_data = self._map_row_with_headers(row, headers)
        else:
            data, extra_data = self._map_row_fixed(row)

        product = self._find_product(data.get('product'))
        if not product:
            raise ValidationError(_('No se encontró el producto: %s') % (data.get('product') or ''))

        product_uom = self._find_uom(data.get('uom'))
        if not product_uom:
            raise ValidationError(_('No se encontró la UDM: %s') % (data.get('uom') or ''))

        quantity = self._to_float(data.get('quantity'), default=1.0)
        price_unit = self._to_float(data.get('price_unit'))
        if price_unit is None:
            raise ValidationError(_('Debes indicar el precio unitario.'))

        planned_dt = self._parse_datetime(data.get('date_planned'))
        if not planned_dt:
            planned_dt = self.purchase_id.date_order or fields.Datetime.now()

        description = (data.get('description') or '').strip()
        if not description:
            description = product.display_name

        vals = {
            'order_id': self.purchase_id.id,
            'product_id': product.id,
            'name': description,
            'product_qty': quantity,
            'product_uom': product_uom.id,
            'price_unit': price_unit,
            'date_planned': fields.Datetime.to_string(planned_dt),
        }

        taxes_value = data.get('taxes')
        if taxes_value not in (None, '', False):
            taxes = self._find_taxes(taxes_value)
            vals['taxes_id'] = [Command.set(taxes.ids)]

        vals.update(self._prepare_extra_field_vals(extra_data))
        return vals

    def _map_row_fixed(self, row):
        padded = list(row) + [None] * max(0, 7 - len(row))
        data = {
            'product': padded[0],
            'description': padded[1],
            'quantity': padded[2],
            'uom': padded[3],
            'price_unit': padded[4],
            'taxes': padded[5],
            'date_planned': padded[6],
        }
        extra = {}
        return data, extra

    def _map_row_with_headers(self, row, headers):
        data = {
            'product': None,
            'description': None,
            'quantity': None,
            'uom': None,
            'price_unit': None,
            'taxes': None,
            'date_planned': None,
        }
        extra = {}

        for index, raw_header in enumerate(headers):
            value = row[index] if index < len(row) else None
            header = (raw_header or '').strip()
            normalized = header.lower()
            matched = False
            for target, aliases in BASE_HEADER_ALIASES.items():
                if normalized in aliases:
                    data[target] = value
                    matched = True
                    break
            if not matched and header:
                extra[header] = value

        # Fallback to fixed columns if essential fields are missing and row looks positional.
        if not data.get('product') and row:
            positional, _extra = self._map_row_fixed(row)
            for key, val in positional.items():
                if data.get(key) in (None, '', False):
                    data[key] = val
        return data, extra

    def _find_product(self, raw_value):
        value = str(raw_value or '').strip()
        if not value:
            return False
        Product = self.env['product.product'].with_company(self.purchase_id.company_id)
        domain_company = ['|', ('company_id', '=', False), ('company_id', '=', self.purchase_id.company_id.id)]

        if self.product_lookup == 'default_code':
            return Product.search(domain_company + [('default_code', '=', value)], limit=1)
        if self.product_lookup == 'barcode':
            return Product.search(domain_company + [('barcode', '=', value)], limit=1)
        if self.product_lookup == 'name':
            product = Product.search(domain_company + [('name', '=', value)], limit=1)
            return product or Product.search(domain_company + [('display_name', '=', value)], limit=1)

        product = Product.search(domain_company + [('default_code', '=', value)], limit=1)
        if product:
            return product
        product = Product.search(domain_company + [('barcode', '=', value)], limit=1)
        if product:
            return product
        product = Product.search(domain_company + [('name', '=', value)], limit=1)
        if product:
            return product
        return Product.search(domain_company + [('display_name', '=', value)], limit=1)

    def _find_uom(self, raw_value):
        value = str(raw_value or '').strip()
        if not value:
            return False
        Uom = self.env['uom.uom']
        uom = Uom.search([('name', '=', value)], limit=1)
        if uom:
            return uom
        return Uom.search([('name', '=ilike', value)], limit=1)

    def _find_taxes(self, raw_value):
        names = [str(v).strip() for v in str(raw_value or '').split(',') if str(v).strip()]
        if not names:
            return self.env['account.tax']
        Tax = self.env['account.tax'].with_company(self.purchase_id.company_id)
        taxes = self.env['account.tax']
        for name in names:
            tax = Tax.search([
                ('type_tax_use', 'in', ['purchase', 'none']),
                '|', ('company_id', '=', False), ('company_id', '=', self.purchase_id.company_id.id),
                ('name', '=', name),
            ], limit=1)
            if not tax:
                raise ValidationError(_('No se encontró el impuesto: %s') % name)
            taxes |= tax
        return taxes

    def _to_float(self, raw_value, default=None):
        if raw_value in (None, '', False):
            return default
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        text = str(raw_value).strip().replace(' ', '')
        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        else:
            text = text.replace(',', '.')
        return float(text)

    def _parse_datetime(self, raw_value):
        if raw_value in (None, '', False):
            return False
        if isinstance(raw_value, datetime):
            return raw_value
        if isinstance(raw_value, date):
            return datetime.combine(raw_value, time.min)
        text = str(raw_value).strip()
        if not text:
            return False
        try:
            parsed = date_parser.parse(text, dayfirst=True)
        except Exception as err:
            raise ValidationError(_('No se pudo interpretar la fecha: %s') % text) from err
        return parsed

    def _prepare_extra_field_vals(self, extra_data):
        vals = {}
        if not extra_data:
            return vals
        line_model = self.env['purchase.order.line']
        for field_name, raw_value in extra_data.items():
            field = line_model._fields.get(field_name)
            if not field or raw_value in (None, '', False):
                continue
            vals[field_name] = self._convert_field_value(field, raw_value)
        return vals

    def _convert_field_value(self, field, raw_value):
        field_type = field.type
        if field_type in ('char', 'text', 'html'):
            return str(raw_value)
        if field_type == 'integer':
            return int(float(raw_value))
        if field_type in ('float', 'monetary'):
            return self._to_float(raw_value, default=0.0)
        if field_type == 'boolean':
            return str(raw_value).strip().lower() in ('1', 'true', 't', 'y', 'yes', 'si', 'sí', 'x')
        if field_type == 'selection':
            value = str(raw_value).strip()
            selection = dict(field.selection)
            if value in selection:
                return value
            inverse = {str(label).lower(): key for key, label in selection.items()}
            lowered = value.lower()
            if lowered in inverse:
                return inverse[lowered]
            raise ValidationError(_('Valor inválido para selección %s: %s') % (field.string, value))
        if field_type == 'many2one':
            relation = self.env[field.comodel_name]
            value = str(raw_value).strip()
            record = relation.search([(relation._rec_name, '=', value)], limit=1)
            if not record:
                record = relation.search([(relation._rec_name, '=ilike', value)], limit=1)
            if not record:
                raise ValidationError(_('No se encontró %s: %s') % (field.string, value))
            return record.id
        if field_type == 'many2many':
            relation = self.env[field.comodel_name]
            values = [v.strip() for v in str(raw_value).split(',') if v.strip()]
            ids = []
            for value in values:
                record = relation.search([(relation._rec_name, '=', value)], limit=1)
                if not record:
                    record = relation.search([(relation._rec_name, '=ilike', value)], limit=1)
                if not record:
                    raise ValidationError(_('No se encontró %s: %s') % (field.string, value))
                ids.append(record.id)
            return [Command.set(ids)]
        if field_type == 'date':
            parsed = self._parse_datetime(raw_value)
            return fields.Date.to_string(parsed.date()) if parsed else False
        if field_type == 'datetime':
            parsed = self._parse_datetime(raw_value)
            return fields.Datetime.to_string(parsed) if parsed else False
        return raw_value

    def _build_result_html(self, created_count, skipped):
        parts = [
            '<div>',
            '<p><strong>%s</strong></p>' % escape(_('Importación finalizada')),
            '<p>%s <strong>%s</strong></p>' % (escape(_('Líneas creadas:')), created_count),
        ]
        if skipped:
            parts.append('<p><strong>%s</strong></p>' % escape(_('Filas omitidas:')))
            parts.append('<ul>')
            for row_number, message in skipped[:100]:
                parts.append('<li>%s %s: %s</li>' % (
                    escape(_('Fila')),
                    row_number,
                    escape(message),
                ))
            parts.append('</ul>')
            if len(skipped) > 100:
                parts.append('<p>%s</p>' % escape(_('Se omitieron líneas adicionales; revisa el archivo para más detalle.')))
        else:
            parts.append('<p>%s</p>' % escape(_('No hubo errores.')))
        parts.append('</div>')
        return ''.join(parts)
